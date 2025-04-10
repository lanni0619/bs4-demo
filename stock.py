# Package
import requests
from bs4 import BeautifulSoup
from apscheduler.schedulers.background import BackgroundScheduler
import openpyxl

# Module
import json
from datetime import datetime
import os
from module.matplotlib_demo import plot_short_selling
from logger import logger

class Stock:
    def __init__(self, stock_code, created_at=None, balance_yest=None, 
                 selling_today=None, return_today=None, balance_today=None, price=None):
        self.stock_code = stock_code
        self.created_at = created_at
        self.balance_yest = balance_yest
        self.selling_today = selling_today
        self.return_today = return_today
        self.balance_today = balance_today
        self.price = price
    
    def toJson(self):
        #  https://stackoverflow.com/questions/7408647/convert-dynamic-python-object-to-json
         return json.dumps(
              self,
              default=lambda o: o.__dict__,
              sort_keys=False,
              indent=4
         )
    
    def crawl_info(self):

        try:
            logger.info(f"crawler function - {self.stock_code} init")
            short_selling_url = 'https://www.twse.com.tw/rwd/zh/marginTrading/TWT93U?response=html'
            price_url = f"https://tw.stock.yahoo.com/quote/{self.stock_code}.TW"

            web1 = requests.get(short_selling_url)
            web2 = requests.get(price_url)

            soup1 = BeautifulSoup(web1.text, "html5lib")
            soup2 = BeautifulSoup(web2.text, "html5lib")

            data_array1 = soup1.find_all('tr', attrs={"align":"center", "style":"font-size:14px;"})
            
            target_classes = [
                "Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-up)", # 漲
                "Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c) C($c-trend-down)", # 跌
                "Fz(32px) Fw(b) Lh(1) Mend(16px) D(f) Ai(c)", # 平盤
                "Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-down)", # 跌停
                "Fz(32px) Fw(b) Lh(1) Mend(16px) C(#fff) Px(6px) Py(2px) Bdrs(4px) Bgc($c-trend-up)", # 漲停
            ]
            data2 = soup2.find('span', class_=lambda x: x and any(cls in x for cls in target_classes))
            price = data2.get_text()

            if price:
                logger.info(f"crawler function - Get price = {price}")
            else:
                logger.error("crawler function - Fail to get price")
                return

            for short_selling in data_array1:
                if short_selling.find('td').get_text() == str(self.stock_code):
                    target_info = short_selling.find_all('td')
                    self.created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.balance_yest = target_info[8].get_text()  # 前日餘額
                    self.selling_today = target_info[9].get_text() # 當日賣出
                    self.return_today = target_info[10].get_text()  # 當日還券
                    self.balance_today = target_info[12].get_text() # 今日餘額
                    self.price = data2.get_text()    
                    logger.info(f"crawler function - Get short selling info")
                    return self
            
            logger.error("crawler function - Fail to get short selling info")
            return self
        
        except requests.exceptions.RequestException as e:
            logger.error(f"crawler function - Network error: {e}")
        except AttributeError as e:
            logger.error(f"crawler function - Parsing error: {e}")
        except Exception as e:
            logger.error(f"crawler function - unexpected error: {e}")
    
    def send_json(self):
        try:
            logger.info(f"send_json - init - stock_number = {self.stock_code}")
            # Preliminary
            if not all(getattr(self, attr) is not None for attr in vars(self)):
                logger.error("send_json - data incomplete")
                return

            info_json = self.toJson()
            url = "https://discord.com/api/webhooks/1356484738029719573/9GNCPHfl7gcz9BpkkO1xYEYqZ9_D2tWd0dx5sZqx3RTN3HgLFLql47TEgWYEsz0Q4x8g"   
            headers = {"Content-Type": "application/json"}
            data = {"content": info_json, "username": "newmanBot"}

            # Send data to discord
            res = requests.post(url, headers = headers, json = data)
            
            # Read the response
            if res.status_code in (200, 204):
                logger.info(f"send_json - success")
            else:
                logger.error(f"send_json - fail")

        except requests.exceptions.RequestException as e:
            logger.error(f"send_json - Network error: {e}")
        except AttributeError as e:
            logger.error(f"send_json - Attribute error: {e}")
        except ValueError as e:
            logger.error(f"send_json - Value error: {e}")
        except Exception as e:
            logger.error(f"send_json - unexpected error: {e}")

    # Save by openpyxl
    def save_to_excel(self):
        try:
            logger.info(f"save_to_excel2 - stock_number = {self.stock_code}")

            # 1) Preliminary
            today = datetime.today()
            root_path = "C:/temp/stock-log"
            filename = os.path.join(root_path, f"{self.stock_code}_{today.strftime('%Y-%m')}.xlsx")
            cols = ["created_at", "balance_yest", "selling_today", "return_today", "balance_today", "price"]

            # 2) Check if data exist ?
            if os.path.exists(filename):
                logger.info(f"save_to_excel2 - file exists, check if duplicate date of record")
                wb = openpyxl.load_workbook(filename, data_only=True) # create work book
                sheet = wb.active                                     # get first sheet when open the xlsx
                max_row = sheet.max_row

                if max_row != 1:                                         # get max_row number
                    str_time = sheet.cell(max_row, 1).value[0:10]        # get last record date (string, yyyy-mm-dd)           
                    # 3) Check if data duplicate ?
                    if str_time == today.strftime("%Y-%m-%d"):            # compare with today
                        logger.info("save_to_excel - Duplicate date of record, stop saving data ")
                        return

            # 4) if file not exists
            else:
                logger.info(f"save_to_excel2 - file not exists, creating ...")
                # create new excel file
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.append(cols)
                wb.save(filename)

            # 5) build new row
            row_arr = []
            for attr in cols:
                value = getattr(self, attr)
                if attr != "created_at" and value is not None:
                    row_arr.append(float(value.replace(",", "")))
                else:
                    row_arr.append(value)

            sheet.append(row_arr)
            wb.save(filename)
            
        except Exception as e:
            print(e)

    def send_chart(self):
        try:
            logger.info(f"send_chart - stock_number = {self.stock_code}")
            # 1) draw new chart
            plot_short_selling(self.stock_code)

            # 2) Path to the JPG file
            today = datetime.today()
            jpg_file_path = f"C:/temp/stock-log/{self.stock_code}_{today.strftime('%Y-%m')}.jpg"  # Replace with your actual file name

            # 3) Discord webhook URL
            url = "https://discord.com/api/webhooks/1356484738029719573/9GNCPHfl7gcz9BpkkO1xYEYqZ9_D2tWd0dx5sZqx3RTN3HgLFLql47TEgWYEsz0Q4x8g"

            # 4) Check if the file exists
            if not os.path.exists(jpg_file_path):
                logger.info(f"send_chart - File not found: {jpg_file_path}")
                return

            # 5) Prepare the file and payload
            with open(jpg_file_path, "rb") as file:
                files = {"file": (os.path.basename(jpg_file_path), file, "image/jpeg")}
                payload = {"username": "newmanBot"}

                # Send the request
                res = requests.post(url, data=payload, files=files)

            # 6) Read the response
            if res.status_code in (200, 204):
                logger.info(f"send_chart - success")
            else:
                logger.info(f"send_chart - fail")
        
        except Exception as e:
            logger.error(f"save_to_excel - error: {e}")

    def schedule_task(self):
        logger.info(f"schedule_task - init - stock_number = {self.stock_code}")
        scheduler = BackgroundScheduler(timezone="Asia/Taipei")
        hour = 21
        min = 30
        sec = 00
        scheduler.add_job(self.crawl_info, 'cron', day_of_week='mon-fri', hour=hour, minute=min, second=sec)
        scheduler.add_job(self.save_to_excel, 'cron', day_of_week='mon-fri', hour=hour, minute=min, second=(sec + 10) % 60)
        scheduler.add_job(self.send_json, 'cron', day_of_week='mon-fri', hour=hour, minute=min, second=(sec + 20) % 60)
        scheduler.add_job(self.send_chart, 'cron', day_of_week='mon-fri', hour=hour, minute=min, second=(sec + 30) % 60)
        scheduler.start()